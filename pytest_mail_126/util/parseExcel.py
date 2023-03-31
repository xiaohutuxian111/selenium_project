"""
@FileName：parseExcel.py
@Author：stone
@Time：2023/3/16 17:43
@Description：
"""
from openpyxl import load_workbook
from config.conf import DATA_PATH


class ParseExcel():

    def __init__(self):
        self.wk = load_workbook(DATA_PATH)
        self.excelFile = DATA_PATH

    # 获取sheet对象
    def getSheetByName(self, sheetName):
        sheet = self.wk.get_sheet_by_name(sheetName)
        return sheet

    # 获取最大的行号
    def getRowNum(self, sheet):
        return sheet.max_row

    #  获取最大的列号
    def getColsNum(self, sheet):
        return sheet.max_column

    # 获取某一行的数据
    def getRowValues(self, sheet, rowNum):
        MaxColsNum = self.getColsNum(sheet)
        rowValues = []
        for colsnum in range(1, MaxColsNum + 1):
            value = sheet.cell(rowNum, colsnum).value
            if value is None:
                value = ""
            rowValues.append(value)
        return tuple(rowValues)

    # 获取某一列的数据
    def getColumnValues(self, sheet, columnNum):
        maxRowNum = self.getRowNum(sheet)
        colValues = []
        for rownum in range(1, maxRowNum + 1):
            value = sheet.cell(rownum, columnNum).value
            if value is None:
                value = ""
            colValues.append(value)
        return tuple(colValues)

    #  获取一个单元格数据
    def getValueofCell(self, sheet, rowNum, columnNum):
        value = sheet.cell(rowNum, columnNum).value
        if value is None:
            value = ""
        return value

    def getAllValuesOfSheet(self, sheet):
        maxRowNum = self.getRowNum(sheet)
        columnNum = self.getColsNum(sheet)
        allVaalue = []
        for row in range(2, maxRowNum + 1):
            rowValues = []
            for column in range(1, columnNum + 1):
                value = sheet.cell(row, column).value
                if value is None:
                    value = ""
                rowValues.append(value)
            allVaalue.append(rowValues)
        return allVaalue


if __name__ == '__main__':
    p = ParseExcel()
    sheet = p.getSheetByName('contact')
    print(p.getRowNum(sheet))
    print(p.getAllValuesOfSheet(sheet))
